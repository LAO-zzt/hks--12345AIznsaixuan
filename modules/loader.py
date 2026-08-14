# -*- coding: utf-8 -*-
"""
模块 1：数据加载与预处理（loader.py）

职责：
- 读取 CSV / xlsx 工单数据（自动识别常见中英文列名）
- 清洗：去空行、按工单编号去重
- 时间字段统一为可计算格式（缺失时从工单编号解析）
- 输出标准字段 DataFrame：order_id / content / title / subject / area / submit_time
"""
import os

import pandas as pd

import config

# 标准字段 -> 可能的中英文列名别名（小写比较）
FIELD_ALIASES = {
    "order_id": ["order_id", "工单编号", "工单号", "编号", "单号", "id", "工单id"],
    "content": ["content", "诉求内容", "工单内容", "内容", "诉求", "事件描述", "问题描述", "描述"],
    "title": ["title", "标题", "工单标题", "事件标题", "事项标题"],
    "subject": ["subject", "涉及主体", "主体", "涉事主体", "涉及对象", "对象", "被诉主体"],
    "area": ["area", "事发区域", "区域", "发生区域", "所属区域", "属地", "地点", "地址"],
    "submit_time": ["submit_time", "提交时间", "受理时间", "创建时间", "时间", "发生时间", "上报时间", "登记时间"],
}

# 输出的标准字段顺序
STANDARD_FIELDS = ["order_id", "content", "title", "subject", "area", "submit_time"]


def _match_columns(df: pd.DataFrame) -> dict:
    """
    自动识别列名，返回 {标准字段: 实际列名}。

    匹配策略：
    1) 精确匹配（忽略大小写/首尾空格）；
    2) 包含匹配（别名出现在列名中，或列名出现在别名中）；
    3) 找不到则留空，交由下游兜底。
    """
    col_map = {}
    lower_cols = {str(c).strip().lower(): c for c in df.columns}

    for field, aliases in FIELD_ALIASES.items():
        matched = None
        # 1) 精确匹配
        for a in aliases:
            if a.lower() in lower_cols:
                matched = lower_cols[a.lower()]
                break
        # 2) 包含匹配
        if matched is None:
            for a in aliases:
                for lc, orig in lower_cols.items():
                    if a.lower() in lc or lc in a.lower():
                        matched = orig
                        break
                if matched is not None:
                    break
        col_map[field] = matched
    return col_map


def _to_datetime(series: pd.Series) -> pd.Series:
    """把时间列统一为 datetime；无法解析的置为 NaT，不抛错。"""
    return pd.to_datetime(series, errors="coerce")


def _parse_time_from_order_id(series: pd.Series) -> pd.Series:
    """
    从工单编号前 6 位解析提交时间（YYMMDD，如 250317→2025-03-17）。

    真实数据常无独立时间列，时间信息内嵌在工单编号中；
    解析失败的置 NaT，不抛错。
    """
    m = series.astype(str).str.extract(r"^(\d{2})(\d{2})(\d{2})")
    parsed = pd.to_datetime(
        m[0] + "-" + m[1] + "-" + m[2], format="%y-%m-%d", errors="coerce")
    return parsed


def load_orders(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    接收原始 DataFrame（由 Streamlit 上传或本地读取），返回标准化后的工单 DataFrame。

    兜底原则：
    - 缺失 order_id：自动生成 AUTO_{行号}；
    - 缺失 content：置空字符串；
    - 缺失 title/subject/area：置空字符串；
    - 缺失/无法解析 submit_time：先尝试从工单编号解析，仍失败置 NaT。
    """
    df = df_raw.copy()
    col_map = _match_columns(df)

    result = pd.DataFrame(index=df.index)
    for field in STANDARD_FIELDS:
        src = col_map.get(field)
        if src is not None and src in df.columns:
            result[field] = df[src]
        else:
            # 位置/缺失兜底
            if field == "order_id":
                result[field] = [f"AUTO_{i}" for i in range(len(df))]
            elif field == "submit_time":
                result[field] = pd.NaT
            else:
                result[field] = ""

    # 统一时间格式
    result["submit_time"] = _to_datetime(result["submit_time"])
    # 时间缺失时尝试从工单编号解析（真实数据常见情形）
    nat_mask = result["submit_time"].isna()
    if nat_mask.any():
        result.loc[nat_mask, "submit_time"] = _parse_time_from_order_id(
            result.loc[nat_mask, "order_id"]).values

    # 文本列转字符串并去首尾空格
    for field in ["order_id", "content", "title", "subject", "area"]:
        result[field] = result[field].fillna("").astype(str).str.strip()

    # 去除完全空行（content 为空视为无效工单）
    result = result[result["content"].str.len() > 0]

    # 按工单编号去重（保留首次出现）
    result = result.drop_duplicates(subset=["order_id"], keep="first")

    result = result.reset_index(drop=True)
    # 附带列映射信息，便于前端展示识别情况
    result.attrs["col_map"] = col_map
    return result


def load_csv_bytes(content: bytes) -> pd.DataFrame:
    """从上传的字节内容读取 CSV，兼容常见编码。"""
    import io
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
        try:
            return pd.read_csv(io.BytesIO(content), encoding=enc)
        except UnicodeDecodeError:
            continue
    # 最后兜底：强制读取
    return pd.read_csv(io.BytesIO(content), encoding="utf-8", errors="ignore")


def read_csv_file(path: str) -> pd.DataFrame:
    """按路径读取 CSV（复用多编码策略）。"""
    with open(path, "rb") as f:
        return load_csv_bytes(f.read())


def read_xlsx_streaming(path: str) -> pd.DataFrame:
    """
    流式读取 xlsx 首个非空 Sheet（openpyxl read_only）。

    相比 pandas.read_excel，内存占用与耗时大幅下降，
    适用于真实数据这类数十 MB、十万行级文件。
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = None
    for name in wb.sheetnames:
        cand = wb[name]
        if cand.max_row and cand.max_row > 1:
            ws = cand
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return pd.DataFrame()
    header = [
        str(h).strip() if h is not None else "col_%d" % i
        for i, h in enumerate(header_row)
    ]
    data = [row for row in rows_iter]
    wb.close()
    return pd.DataFrame(data, columns=header)


CACHE_DIR = os.path.join(config.OUTPUT_DIR, "cache")


def load_orders_cached(path: str) -> pd.DataFrame:
    """
    统一文件加载入口（CSV/xlsx）+ 解析缓存。

    以 文件名+修改时间+大小 为缓存键，首次解析后落地 pickle，
    之后秒级加载，保证大数据现场演示不卡在读取阶段。
    """
    os.makedirs(CACHE_DIR, exist_ok=True)
    st = os.stat(path)
    stem = os.path.splitext(os.path.basename(path))[0]
    cache_path = os.path.join(
        CACHE_DIR, "%s_%d_%d.pkl" % (stem, int(st.st_mtime), st.st_size))
    if os.path.exists(cache_path):
        try:
            return pd.read_pickle(cache_path)
        except Exception:
            pass

    if path.lower().endswith((".xlsx", ".xlsm")):
        df_raw = read_xlsx_streaming(path)
    else:
        df_raw = read_csv_file(path)

    df = load_orders(df_raw)
    try:
        df.to_pickle(cache_path)
    except Exception:
        pass
    return df
